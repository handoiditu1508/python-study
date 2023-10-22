import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook("tutorial_1/excel_automation/transactions.xlsx")
sheet = wb["Sheet1"]

print("-" * 10, "testing sheet object", "-" * 10)

# cell = sheet["a1"]
cell = sheet.cell(1, 1)
print("first cell value: ", cell.value)
print("sheet max row: ", sheet.max_row)

print("-" * 10, "create new excel file out of old one", "-" * 10)

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price

wb.save("tutorial_1/excel_automation/transactions2.xlsx")

print("-" * 10, "create chart", "-" * 10)

values = Reference(
    sheet,
    min_row=2,
    max_row=sheet.max_row,
    min_col=4,
    max_col=4,
)

chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, "e2")

wb.save("tutorial_1/excel_automation/transactions3.xlsx")
